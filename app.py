import streamlit as st
import pandas as pd
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 설정 ────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="AI 데이터 자동 취합 시스템",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

DEPARTMENTS = [
    "MD팀",
    "스탭부서",
    "점포운영팀",
    "마케팅팀",
    "프로모션팀",
    "광고팀",
    "상품기획팀",
    "영업팀",
    "VMD팀",
    "행사기획팀",
]

REQUIRED_COLUMNS = ["브랜드명", "할인율(%)", "참여점포수", "상품구성", "비고"]

# ── 세션 상태 초기화 ─────────────────────────────────────────────────────────
if "master_df" not in st.session_state:
    st.session_state.master_df = pd.DataFrame(columns=["부서"] + REQUIRED_COLUMNS + ["업로드시각", "버전"])
if "upload_log" not in st.session_state:
    st.session_state.upload_log = []
if "change_log" not in st.session_state:
    st.session_state.change_log = []


# ── 유틸 함수 ────────────────────────────────────────────────────────────────
def validate_columns(df: pd.DataFrame) -> list[str]:
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    return missing


def detect_changes(old_df: pd.DataFrame, new_df: pd.DataFrame, dept: str) -> list[str]:
    changes = []
    old_dept = old_df[old_df["부서"] == dept].copy()
    if old_dept.empty:
        for _, row in new_df.iterrows():
            changes.append(f"[신규] {dept} — '{row['브랜드명']}' 데이터 추가")
        return changes

    old_dept = old_dept.set_index("브랜드명")
    new_idx = new_df.set_index("브랜드명")

    added = set(new_idx.index) - set(old_dept.index)
    removed = set(old_dept.index) - set(new_idx.index)

    for brand in added:
        changes.append(f"[추가] {dept} — '{brand}' 브랜드 신규 등록")
    for brand in removed:
        changes.append(f"[삭제] {dept} — '{brand}' 브랜드 제거")

    for brand in set(new_idx.index) & set(old_dept.index):
        for col in ["할인율(%)", "참여점포수", "상품구성"]:
            if col in new_idx.columns and col in old_dept.columns:
                old_val = old_dept.loc[brand, col]
                new_val = new_idx.loc[brand, col]
                if str(old_val) != str(new_val):
                    changes.append(f"[변경] {dept} — '{brand}' {col}: {old_val} → {new_val}")
    return changes


def merge_data(master: pd.DataFrame, new_df: pd.DataFrame, dept: str, version: int) -> pd.DataFrame:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    new_df = new_df.copy()
    new_df["부서"] = dept
    new_df["업로드시각"] = now
    new_df["버전"] = version

    master = master[master["부서"] != dept].copy()
    master = pd.concat([master, new_df[["부서"] + REQUIRED_COLUMNS + ["업로드시각", "버전"]]], ignore_index=True)
    return master


def check_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    return df[df.duplicated(subset=["부서", "브랜드명"], keep=False)]


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "마스터시트"

    header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    dept_colors = {
        "MD팀": "E8F4FD", "스탭부서": "E8F8E8", "점포운영팀": "FFF8E1",
        "마케팅팀": "FCE4EC", "프로모션팀": "F3E5F5", "광고팀": "E0F7FA",
        "상품기획팀": "FFF3E0", "영업팀": "E8EAF6", "VMD팀": "F9FBE7", "행사기획팀": "FAFAFA",
    }

    cols = ["부서"] + REQUIRED_COLUMNS + ["업로드시각", "버전"]
    col_widths = [14, 20, 12, 14, 20, 30, 20, 8]

    for ci, (col, width) in enumerate(zip(cols, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[1].height = 22

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        dept = str(row.get("부서", ""))
        fill_color = dept_colors.get(dept, "FFFFFF")
        fill = PatternFill("solid", start_color=fill_color, end_color=fill_color)
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col, ""))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = fill
            cell.border = border

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def sample_excel_bytes(dept: str) -> bytes:
    df = pd.DataFrame({
        "브랜드명": ["브랜드A", "브랜드B", "브랜드C"],
        "할인율(%)": [10, 15, 20],
        "참여점포수": [5, 8, 12],
        "상품구성": ["기본형", "프리미엄", "스탠다드"],
        "비고": ["", "신규", ""],
    })
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="데이터입력")
    return buf.getvalue()


# ── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-title {
        font-size: 1.8rem; font-weight: 800; color: #1F4E79;
        border-bottom: 3px solid #2196F3; padding-bottom: 8px; margin-bottom: 4px;
    }
    .sub-title { font-size: 0.95rem; color: #666; margin-bottom: 24px; }
    .kpi-card {
        background: linear-gradient(135deg, #f0f8ff, #e3f2fd);
        border-left: 4px solid #2196F3; border-radius: 8px;
        padding: 16px 20px; margin-bottom: 8px;
    }
    .kpi-label { font-size: 0.78rem; color: #666; font-weight: 600; text-transform: uppercase; }
    .kpi-value { font-size: 1.5rem; font-weight: 800; color: #1F4E79; }
    .kpi-sub { font-size: 0.75rem; color: #4CAF50; font-weight: 600; }
    .change-item {
        background: #fff; border-left: 3px solid #4CAF50; border-radius: 4px;
        padding: 6px 12px; margin: 4px 0; font-size: 0.85rem; color: #333;
    }
    .change-add { border-left-color: #2196F3; }
    .change-del { border-left-color: #F44336; }
    .change-mod { border-left-color: #FF9800; }
    .section-header {
        font-size: 1.05rem; font-weight: 700; color: #1F4E79;
        margin: 20px 0 10px 0; padding-bottom: 4px;
        border-bottom: 2px solid #e0e0e0;
    }
    .upload-success {
        background: #E8F5E9; border: 1px solid #4CAF50; border-radius: 6px;
        padding: 10px 16px; color: #2E7D32; font-weight: 600;
    }
    .upload-error {
        background: #FFEBEE; border: 1px solid #F44336; border-radius: 6px;
        padding: 10px 16px; color: #C62828; font-weight: 600;
    }
</style>
""", unsafe_allow_html=True)

# ── 헤더 ────────────────────────────────────────────────────────────────────
st.markdown('<div class="main-title">📊 AI 기반 엑셀 데이터 자동 취합 시스템</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">업로드만 하면 자동으로 취합되는 Zero-Touch 데이터 운영 구조</div>', unsafe_allow_html=True)

# ── 사이드바 ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📁 파일 업로드")
    st.markdown("---")

    dept = st.selectbox("부서 선택", DEPARTMENTS)

    st.markdown("**표준 양식 다운로드**")
    st.download_button(
        "⬇️ 샘플 엑셀 다운로드",
        data=sample_excel_bytes(dept),
        file_name=f"표준양식_{dept}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.markdown("---")
    uploaded = st.file_uploader("엑셀 파일 업로드 (.xlsx)", type=["xlsx", "xls"])

    if st.button("🚀 데이터 취합 실행", use_container_width=True, type="primary"):
        if uploaded is None:
            st.error("파일을 먼저 업로드해주세요.")
        else:
            try:
                df_new = pd.read_excel(uploaded)
                missing = validate_columns(df_new)
                if missing:
                    st.markdown(f'<div class="upload-error">❌ 필수 컬럼 누락: {", ".join(missing)}</div>', unsafe_allow_html=True)
                else:
                    changes = detect_changes(st.session_state.master_df, df_new, dept)

                    dept_versions = st.session_state.master_df[
                        st.session_state.master_df["부서"] == dept
                    ]["버전"].tolist()
                    version = (max(dept_versions) + 1) if dept_versions else 1

                    st.session_state.master_df = merge_data(
                        st.session_state.master_df, df_new, dept, version
                    )

                    now_str = datetime.now().strftime("%H:%M:%S")
                    st.session_state.upload_log.append({
                        "시각": now_str,
                        "부서": dept,
                        "행수": len(df_new),
                        "버전": version,
                        "변경수": len(changes),
                    })
                    st.session_state.change_log = changes + st.session_state.change_log

                    st.markdown(
                        f'<div class="upload-success">✅ {dept} 데이터 취합 완료 (v{version}, {len(df_new)}건)</div>',
                        unsafe_allow_html=True,
                    )
                    if changes:
                        st.info(f"변경사항 {len(changes)}건 감지됨")
            except Exception as e:
                st.error(f"처리 오류: {e}")

    if st.button("🗑️ 전체 데이터 초기화", use_container_width=True):
        st.session_state.master_df = pd.DataFrame(columns=["부서"] + REQUIRED_COLUMNS + ["업로드시각", "버전"])
        st.session_state.upload_log = []
        st.session_state.change_log = []
        st.success("초기화 완료")

    st.markdown("---")
    st.markdown("**📋 필수 컬럼 안내**")
    for col in REQUIRED_COLUMNS:
        st.markdown(f"- `{col}`")

# ── 메인 콘텐츠 ──────────────────────────────────────────────────────────────
master = st.session_state.master_df

total_brands = len(master["브랜드명"].unique()) if not master.empty else 0
total_dept = len(master["부서"].unique()) if not master.empty else 0
total_stores = int(master["참여점포수"].sum()) if not master.empty and "참여점포수" in master.columns else 0
dupes = check_duplicates(master)

col1, col2, col3, col4 = st.columns(4)
with col1:
    st.markdown(f"""<div class="kpi-card">
        <div class="kpi-label">참여 부서</div>
        <div class="kpi-value">{total_dept} / {len(DEPARTMENTS)}</div>
        <div class="kpi-sub">취합 부서 수</div></div>""", unsafe_allow_html=True)
with col2:
    st.markdown(f"""<div class="kpi-card">
        <div class="kpi-label">총 브랜드 수</div>
        <div class="kpi-value">{total_brands}</div>
        <div class="kpi-sub">고유 브랜드 기준</div></div>""", unsafe_allow_html=True)
with col3:
    st.markdown(f"""<div class="kpi-card">
        <div class="kpi-label">총 참여 점포</div>
        <div class="kpi-value">{total_stores}</div>
        <div class="kpi-sub">전체 합산 기준</div></div>""", unsafe_allow_html=True)
with col4:
    dup_color = "#F44336" if not dupes.empty else "#4CAF50"
    dup_label = f"⚠️ {len(dupes)}건" if not dupes.empty else "✅ 없음"
    st.markdown(f"""<div class="kpi-card">
        <div class="kpi-label">중복 데이터</div>
        <div class="kpi-value" style="color:{dup_color}">{dup_label}</div>
        <div class="kpi-sub">오류 탐지 결과</div></div>""", unsafe_allow_html=True)

st.markdown("---")

tab1, tab2, tab3, tab4 = st.tabs(["📋 마스터 시트", "🔔 변경 알림", "📂 업로드 이력", "📊 부서별 현황"])

with tab1:
    st.markdown('<div class="section-header">통합 마스터 시트 (실시간 최신)</div>', unsafe_allow_html=True)

    if master.empty:
        st.info("아직 업로드된 데이터가 없습니다. 왼쪽 사이드바에서 파일을 업로드해주세요.")
    else:
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            dept_filter = st.multiselect("부서 필터", options=master["부서"].unique().tolist(), default=[])
        with col_f2:
            brand_search = st.text_input("브랜드 검색", placeholder="브랜드명 입력...")

        filtered = master.copy()
        if dept_filter:
            filtered = filtered[filtered["부서"].isin(dept_filter)]
        if brand_search:
            filtered = filtered[filtered["브랜드명"].str.contains(brand_search, case=False, na=False)]

        st.dataframe(
            filtered.reset_index(drop=True),
            use_container_width=True,
            height=400,
            column_config={
                "할인율(%)": st.column_config.NumberColumn("할인율(%)", format="%d%%"),
                "참여점포수": st.column_config.NumberColumn("참여점포수", format="%d개"),
                "버전": st.column_config.NumberColumn("버전", format="v%d"),
            }
        )

        st.markdown(f"**총 {len(filtered)}건** | 마지막 업데이트: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

        excel_data = to_excel_bytes(filtered)
        st.download_button(
            "⬇️ 마스터시트 다운로드 (.xlsx)",
            data=excel_data,
            file_name=f"마스터시트_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        if not dupes.empty:
            st.warning(f"⚠️ 중복 데이터 {len(dupes)}건 발견")
            st.dataframe(dupes, use_container_width=True)

with tab2:
    st.markdown('<div class="section-header">자동 변경 알림 (최근 순)</div>', unsafe_allow_html=True)

    if not st.session_state.change_log:
        st.info("변경 이력이 없습니다.")
    else:
        for item in st.session_state.change_log[:30]:
            if "[추가]" in item or "[신규]" in item:
                css = "change-item change-add"
            elif "[삭제]" in item:
                css = "change-item change-del"
            else:
                css = "change-item change-mod"
            st.markdown(f'<div class="{css}">🔔 {item}</div>', unsafe_allow_html=True)

with tab3:
    st.markdown('<div class="section-header">업로드 이력</div>', unsafe_allow_html=True)

    if not st.session_state.upload_log:
        st.info("업로드 이력이 없습니다.")
    else:
        log_df = pd.DataFrame(st.session_state.upload_log)
        st.dataframe(log_df, use_container_width=True)

with tab4:
    st.markdown('<div class="section-header">부서별 데이터 현황</div>', unsafe_allow_html=True)

    submitted = set(master["부서"].unique()) if not master.empty else set()
    pending = set(DEPARTMENTS) - submitted

    col_s, col_p = st.columns(2)
    with col_s:
        st.markdown("**✅ 제출 완료 부서**")
        for d in sorted(submitted):
            cnt = len(master[master["부서"] == d])
            ver = int(master[master["부서"] == d]["버전"].max())
            st.markdown(f"- {d} — {cnt}건 (v{ver})")
    with col_p:
        st.markdown("**⏳ 미제출 부서**")
        if pending:
            for d in sorted(pending):
                st.markdown(f"- {d}")
        else:
            st.success("모든 부서 제출 완료!")

    if not master.empty:
        st.markdown("---")
        st.markdown("**부서별 브랜드 수**")
        dept_summary = (
            master.groupby("부서")
            .agg(브랜드수=("브랜드명", "count"), 평균할인율=("할인율(%)", "mean"), 총점포수=("참여점포수", "sum"))
            .reset_index()
        )
        dept_summary["평균할인율"] = dept_summary["평균할인율"].round(1)
        st.dataframe(dept_summary, use_container_width=True)
